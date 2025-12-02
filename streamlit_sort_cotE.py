import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def sort_dataframe_by_column_e(df: pd.DataFrame) -> pd.DataFrame:
    """Sắp xếp DataFrame theo cột E (cột thứ 5) theo thứ tự tăng dần."""
    if len(df.columns) < 5:
        raise ValueError("Bảng không có đủ 5 cột để xác định cột E.")

    col_e_name = df.columns[4]
    return df.sort_values(by=col_e_name, ascending=True)


def auto_format_workbook(wb) -> None:
    """Căn lề, tự điều chỉnh độ rộng cột và wrap text cho từng ô."""
    ws = wb.active

    for column_cells in ws.columns:
        first_cell = column_cells[0]
        column_letter = get_column_letter(first_cell.column)

        max_length = 0
        for cell in column_cells:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))

        ws.column_dimensions[column_letter].width = max_length + 2


def build_downloadable_workbook(df: pd.DataFrame) -> BytesIO:
    """Ghi DataFrame vào workbook, định dạng và trả về buffer BytesIO."""
    temp_buffer = BytesIO()
    with pd.ExcelWriter(temp_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    temp_buffer.seek(0)
    wb = load_workbook(temp_buffer)
    auto_format_workbook(wb)

    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


def main() -> None:
    st.set_page_config(page_title="Sắp xếp cột E", layout="wide")
    st.title("Sắp xếp dữ liệu Excel theo cột E (tăng dần)")
    st.write(
        "Tải file Excel lên, hệ thống sẽ sắp xếp dữ liệu theo cột E và trả về file đã canh chỉnh."
    )

    uploaded_file = st.file_uploader("Chọn file Excel", type=["xlsx", "xlsm", "xls"])

    if uploaded_file is None:
        st.info("Vui lòng tải file Excel để bắt đầu.")
        return

    try:
        df = pd.read_excel(uploaded_file)
    except Exception as exc:
        st.error(f"Không thể đọc file Excel: {exc}")
        return

    try:
        sorted_df = sort_dataframe_by_column_e(df)
    except ValueError as exc:
        st.error(str(exc))
        return

    st.success("Đã sắp xếp dữ liệu theo cột E (tăng dần).")
    st.dataframe(sorted_df, use_container_width=True)

    formatted_workbook = build_downloadable_workbook(sorted_df)

    st.download_button(
        label="Tải file Excel đã sắp xếp",
        data=formatted_workbook,
        file_name="tonghop_sapxep_cotE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()


