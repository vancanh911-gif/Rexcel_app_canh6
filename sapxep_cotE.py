import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def sort_and_format_by_column_e(
    input_path: str = "tonghop.xlsx",
    output_path: str = "tonghop_sapxep_cotE.xlsx",
) -> None:
    """
    Sắp xếp dữ liệu theo cột E (từ tăng dần) và canh chỉnh dòng/cột cơ bản.

    - Cột E tương ứng với cột thứ 5 trong bảng (index = 4).
    - File đầu ra giữ nguyên tên cột, không ghi chỉ số dòng.
    - Sau khi ghi, file được canh chỉnh căn lề trái, giữa theo chiều dọc
      và tự điều chỉnh độ rộng cột dựa trên nội dung.
    """
    df = pd.read_excel(input_path)

    if len(df.columns) < 5:
        raise ValueError("Bảng không có đủ 5 cột để xác định cột E.")

    col_e_name = df.columns[4]

    df_sorted = df.sort_values(by=col_e_name, ascending=True)
    df_sorted.to_excel(output_path, index=False)

    _auto_format_excel(output_path)


def _auto_format_excel(file_path: str) -> None:
    """Căn chỉnh cơ bản: wrap text, căn trái và chỉnh rộng cột."""
    wb = load_workbook(file_path)
    ws = wb.active

    for column_cells in ws.columns:
        first_cell = column_cells[0]
        column_letter = get_column_letter(first_cell.column)

        max_length = 0
        for cell in column_cells:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(file_path)


if __name__ == "__main__":
    sort_and_format_by_column_e()


